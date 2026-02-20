import pandas as pd
from database.transaction import readonly
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


class RelatorioService:

    # ==========================================================
    # FUNÇÕES AUXILIARES
    # ==========================================================

    @staticmethod
    def segundos_para_hhmmss(segundos: int) -> str:
        horas = segundos // 3600
        minutos = (segundos % 3600) // 60
        seg = segundos % 60
        return f"{horas:02d}:{minutos:02d}:{seg:02d}"

    @staticmethod
    def segundos_para_pace(segundos: int) -> str:
        minutos = segundos // 60
        seg = segundos % 60
        return f"{int(minutos):02d}:{int(seg):02d} /km"

    # ==========================================================
    # RELATÓRIO MENSAL COMPLETO
    # ==========================================================

    @readonly
    def gerar_relatorio_mensal(self, mes: str, *, conn=None) -> str:

        # ==========================
        # SHEET 1 - RESUMO MENSAL
        # ==========================

        query_resumo = """
        SELECT
            nome,
            total_treinos,
            km_total,
            tempo_total_segundos,
            pace_medio_segundos
        FROM corridas_mensais
        WHERE TO_CHAR(mes, 'YYYY-MM') = %s
        ORDER BY km_total DESC
        """

        # df_resumo = pd.read_sql(query_resumo, conn, params=(mes,))
        cursor = conn.cursor()
        cursor.execute(query_resumo, (mes,))
        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        df_resumo = pd.DataFrame(rows, columns=columns)
        cursor.close()

        if df_resumo.empty:
            raise ValueError("Nenhum registro encontrado para o mês informado.")

        df_resumo["Tempo Total"] = df_resumo["tempo_total_segundos"].apply(
            self.segundos_para_hhmmss
        )

        df_resumo["Pace Médio"] = df_resumo["pace_medio_segundos"].apply(
            self.segundos_para_pace
        )

        df_resumo_final = df_resumo[
            ["nome", "total_treinos", "km_total", "Tempo Total", "Pace Médio"]
        ].copy()

        df_resumo_final.rename(columns={
            "nome": "Atleta",
            "total_treinos": "Total Treinos",
            "km_total": "KM Total"
        }, inplace=True)

        # ==========================
        # SHEET 2 - TREINOS DETALHADOS
        # ==========================

        query_detalhado = """
        SELECT
            u.nome,
            c.data_corrida,
            ROUND(c.distancia_metros / 1000.0, 2) AS km,
            c.tempo_segundos,
            c.pace_segundos
        FROM corridas c
        JOIN usuarios u ON u.telegram_id = c.telegram_id
        WHERE TO_CHAR(c.data_corrida, 'YYYY-MM') = %s
        ORDER BY c.data_corrida DESC
        """

        # df_detalhado = pd.read_sql(query_detalhado, conn, params=(mes,))
        cursor = conn.cursor()
        cursor.execute(query_detalhado, (mes,))
        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        df_detalhado = pd.DataFrame(rows, columns=columns)
        cursor.close()

        df_detalhado["Tempo"] = df_detalhado["tempo_segundos"].apply(
            self.segundos_para_hhmmss
        )

        df_detalhado["Pace"] = df_detalhado["pace_segundos"].apply(
            self.segundos_para_pace
        )

        df_detalhado_final = df_detalhado[
            ["nome", "data_corrida", "km", "Tempo", "Pace"]
        ].copy()

        df_detalhado_final.rename(columns={
            "nome": "Atleta",
            "data_corrida": "Data",
            "km": "KM"
        }, inplace=True)

        # ==========================
        # EXPORTAÇÃO
        # ==========================

        arquivo = f"relatorio_corridas_{mes}.xlsx"

        with pd.ExcelWriter(arquivo, engine="openpyxl") as writer:
            df_resumo_final.to_excel(writer, sheet_name="Resumo Mensal", index=False)
            df_detalhado_final.to_excel(writer, sheet_name="Treinos Detalhados", index=False)

        # ==========================
        # FORMATAÇÃO PROFISSIONAL
        # ==========================

        wb = load_workbook(arquivo)

        for sheet in wb.sheetnames:
            ws = wb[sheet]

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            header_fill = PatternFill(
                start_color="1F4E78",
                end_color="1F4E78",
                fill_type="solid"
            )

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for column_cells in ws.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        wb.save(arquivo)

        return arquivo